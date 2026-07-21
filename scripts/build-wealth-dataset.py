#!/usr/bin/env python3
"""Build lib/dgc/wealth-data/dataset.json from downloadable primary sources.

Every populated cell is derived from a source file that anyone can download and
check, or from a specific printed table cited per cell:

  * Federal Reserve Distributional Financial Accounts (1989Q3+): wealth shares,
    total household net worth, household counts for 1990-2025.
    https://www.federalreserve.gov/releases/z1/dataviz/download/zips/dfa.zip
  * Piketty-Saez-Zucman Appendix Tables I (Table TB1): total household net
    worth, nominal USD, 1913-2021 - used for 1920-1985.
    https://gabriel-zucman.eu/files/PSZ2022AppendixTablesI(Aggreg).xlsx
  * Piketty-Saez-Zucman Appendix Tables II (Table TE1): wealth shares by
    percentile group, 1913+ - used for 1920-1985. Bottom-50% / 50-90% detail
    only exists from 1962, so those cells are N/A before 1965.
    https://gabriel-zucman.eu/files/PSZ2022AppendixTablesII(Distrib).xlsx
  * Census HH-1 households table (1940+) and Historical Statistics of the U.S.
    (pre-1940 decennial "families" counts) for household counts.
    https://www2.census.gov/programs-surveys/demo/tables/families/time-series/households/hh1.xls
  * Census historical national population estimates (popclockest.txt) plus
    decennial census counts for population.
    https://www2.census.gov/programs-surveys/popest/tables/1900-1980/national/totals/popclockest.txt
  * Wolff (NBER w28383), Table 1 Panel A row 3a for percent of households with
    zero or negative net worth (SCF survey years mapped to nearest grid year).

Values that cannot be reproduced from any of these sources are emitted as N/A
rather than estimated.

Requirements: python3 with openpyxl and xlrd (pip install openpyxl xlrd).
Downloads are cached in /tmp/wealth-research.
"""

from __future__ import annotations

import csv
import io
import json
import re
import urllib.request
import zipfile
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "lib/dgc/wealth-data/dataset.json"
CACHE = Path("/tmp/wealth-research")
ACCESS = date.today().isoformat()

GRID_YEARS = list(range(1920, 2026, 5))

# ---------------------------------------------------------------------------
# Source directory (also embedded in the dataset for the /dgc/data page)
# ---------------------------------------------------------------------------

SOURCES: dict[str, dict[str, Any]] = {
    "fed-dfa": {
        "id": "fed-dfa",
        "name": "Federal Reserve Distributional Financial Accounts",
        "url": "https://www.federalreserve.gov/releases/z1/dataviz/dfa/",
        "dataFileUrl": "https://www.federalreserve.gov/releases/z1/dataviz/download/zips/dfa.zip",
        "coverage": "1989:Q3 onward, quarterly",
        "description": (
            "Official quarterly estimates of U.S. household wealth by wealth percentile group. "
            "Used for all wealth shares, total net worth, and household counts from 1990."
        ),
    },
    "psz-agg": {
        "id": "psz-agg",
        "name": "Piketty-Saez-Zucman Appendix Tables I (aggregates)",
        "url": "https://gabriel-zucman.eu/usdina/",
        "dataFileUrl": "https://gabriel-zucman.eu/files/PSZ2022AppendixTablesI(Aggreg).xlsx",
        "coverage": "1913-2021, annual",
        "description": (
            "Table TB1 'Household wealth' column [1] (net household wealth, billions of current USD). "
            "Used for total household wealth 1920-1985."
        ),
    },
    "psz-dist": {
        "id": "psz-dist",
        "name": "Piketty-Saez-Zucman Appendix Tables II (distributions)",
        "url": "https://gabriel-zucman.eu/usdina/",
        "dataFileUrl": "https://gabriel-zucman.eu/files/PSZ2022AppendixTablesII(Distrib).xlsx",
        "coverage": "Top shares 1913+; bottom-50%/middle-40% detail from 1962",
        "description": (
            "Table TE1 'Shares of total household wealth' (equal-split adults). "
            "Used for wealth shares 1920-1985. Bottom-half detail does not exist before 1962."
        ),
    },
    "census-hh": {
        "id": "census-hh",
        "name": "U.S. Census Bureau - Historical Households (HH-1)",
        "url": "https://www.census.gov/data/tables/time-series/demo/families/households.html",
        "dataFileUrl": "https://www2.census.gov/programs-surveys/demo/tables/families/time-series/households/hh1.xls",
        "coverage": "1940 onward (no surveys 1941-1946)",
        "description": "Table HH-1 'Households by Type', total households. Used for household counts from 1940.",
    },
    "hsus": {
        "id": "hsus",
        "name": "Historical Statistics of the U.S., Colonial Times to 1970",
        "url": "https://www.census.gov/library/publications/1975/compendia/hist_stats_colonial-1970.html",
        "coverage": "Decennial census counts to 1970",
        "description": (
            "Bicentennial compendium of census series. Pre-1940 decennial household ('families') counts."
        ),
    },
    "census-pop-hist": {
        "id": "census-pop-hist",
        "name": "U.S. Census Bureau - Historical National Population Estimates",
        "url": "https://www.census.gov/data/tables/time-series/demo/popest/pre-1980-national.html",
        "dataFileUrl": "https://www2.census.gov/programs-surveys/popest/tables/1900-1980/national/totals/popclockest.txt",
        "coverage": "July 1, 1900 to July 1, 1999",
        "description": "Official July 1 national population estimates. Used for non-census grid years before 2000.",
    },
    "census-pop-modern": {
        "id": "census-pop-modern",
        "name": "U.S. Census Bureau - Decennial counts and population estimates",
        "url": "https://www.census.gov/programs-surveys/popest/data/tables.html",
        "coverage": "Decennial censuses; vintage estimates 2000 onward",
        "description": "Decennial census resident-population counts and post-2000 vintage estimates.",
    },
    "wolff": {
        "id": "wolff",
        "name": "Edward N. Wolff (NBER Working Paper 28383)",
        "url": "https://www.nber.org/papers/w28383",
        "dataFileUrl": "https://www.nber.org/system/files/working_papers/w28383/w28383.pdf",
        "coverage": "SCF survey years 1962-2019",
        "description": (
            "Table 1 Panel A row 3a: percent of U.S. households with zero or negative net worth. "
            "Survey years are mapped to the nearest five-year grid point."
        ),
    },
}


def src(source_id: str, table: str | None = None) -> dict[str, Any]:
    entry = SOURCES[source_id]
    citation: dict[str, Any] = {
        "name": entry["name"],
        "url": entry["url"],
        "tableOrSeries": table,
        "accessedAt": ACCESS,
    }
    if entry.get("dataFileUrl"):
        citation["dataFileUrl"] = entry["dataFileUrl"]
    return citation


def cell(
    value: float | None,
    unit: str,
    confidence: str,
    source: dict | None,
    *,
    observation_year: int | None = None,
    is_interpolated: bool = False,
    definition: str | None = None,
    methodology_note: str | None = None,
) -> dict[str, Any]:
    return {
        "value": value,
        "unit": unit,
        "observationYear": observation_year,
        "isInterpolated": is_interpolated,
        "confidence": confidence,
        "definition": definition,
        "methodologyNote": methodology_note,
        "source": source,
    }


def na_cell(note: str) -> dict[str, Any]:
    return cell(None, "percent", "na", None, methodology_note=note)


def fetch(url: str, filename: str) -> bytes:
    path = CACHE / filename
    if path.exists():
        return path.read_bytes()
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (DGC wealth-data builder)"})
    with urllib.request.urlopen(req) as resp:
        data = resp.read()
    CACHE.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)
    return data


# ---------------------------------------------------------------------------
# Federal Reserve DFA (Tier 1, 1990-2025)
# ---------------------------------------------------------------------------

DFA_CATS = ["TopPt1", "RemainingTop1", "Next9", "Next40", "Bottom50"]
SHARE_KEYS = [
    "shareTop01Pct",
    "share99to999Pct",
    "share90to99Pct",
    "share50to90Pct",
    "shareBottom50Pct",
]


def load_dfa() -> tuple[dict, dict]:
    data = fetch(SOURCES["fed-dfa"]["dataFileUrl"], "dfa.zip")
    zf = zipfile.ZipFile(io.BytesIO(data))
    shares: dict[str, dict[str, float]] = {}
    levels: dict[str, dict[str, dict[str, float]]] = {}
    for name in ("dfa-networth-shares-detail.csv", "dfa-networth-levels-detail.csv"):
        raw = zf.read(name).decode("utf-8")
        for row in csv.DictReader(raw.splitlines()):
            date_key = row["Date"]
            cat = row["Category"]
            if name.startswith("dfa-networth-shares"):
                shares.setdefault(date_key, {})[cat] = float(row["Net worth"])
            else:
                levels.setdefault(date_key, {})[cat] = {
                    "net_worth": float(row["Net worth"]),
                    "household_count": float(row["Household count"]),
                }
    return shares, levels


def pick_quarter(shares: dict, year: int) -> str | None:
    for q in (3, 4, 2, 1):
        key = f"{year}:Q{q}"
        if key in shares:
            return key
    return None


# ---------------------------------------------------------------------------
# PSZ workbooks (Tier 2/3, 1920-1985)
# ---------------------------------------------------------------------------


def load_psz_total_wealth() -> dict[int, float]:
    """Table TB1 column [1]: net household wealth, bn current USD -> millions."""
    import openpyxl

    data = fetch(SOURCES["psz-agg"]["dataFileUrl"], "psz-tables1.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["TB1"]
    wealth: dict[int, float] = {}
    for row in ws.iter_rows(min_row=10, max_col=2, values_only=True):
        year, value = row[0], row[1]
        if isinstance(year, (int, float)) and value is not None:
            wealth[int(year)] = round(float(value) * 1000)
    return wealth


def load_psz_shares() -> dict[int, dict[str, float | None]]:
    """Table TE1: shares of total household wealth (fractions).

    Columns: A=year, C=Bottom 50%, D=Middle 40%, E=Top 10%, G=Top 1%, I=Top 0.1%.
    Bottom 50% / Middle 40% are only populated from 1962.
    """
    import openpyxl

    data = fetch(SOURCES["psz-dist"]["dataFileUrl"], "psz-tables2.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["TE1"]
    shares: dict[int, dict[str, float | None]] = {}
    for row in ws.iter_rows(min_row=10, max_col=9, values_only=True):
        year = row[0]
        if not isinstance(year, (int, float)):
            continue
        top10, top1, top01 = row[4], row[6], row[8]
        if top10 is None or top1 is None or top01 is None:
            continue
        bot50, mid40 = row[2], row[3]
        shares[int(year)] = {
            "shareTop01Pct": round(top01 * 100, 1),
            "share99to999Pct": round((top1 - top01) * 100, 1),
            "share90to99Pct": round((top10 - top1) * 100, 1),
            "share50to90Pct": round(mid40 * 100, 1) if mid40 is not None else None,
            "shareBottom50Pct": round(bot50 * 100, 1) if bot50 is not None else None,
        }
    return shares


# ---------------------------------------------------------------------------
# Census households
# ---------------------------------------------------------------------------


def load_census_households() -> dict[int, int]:
    """Table HH-1 total households (thousands -> count). First entry per year wins
    (revisions like '2021r' appear above the original)."""
    import xlrd

    data = fetch(SOURCES["census-hh"]["dataFileUrl"], "hh1.xls")
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    households: dict[int, int] = {}
    for r in range(11, ws.nrows):
        raw_year = ws.cell_value(r, 0)
        total = ws.cell_value(r, 1)
        if isinstance(raw_year, float):
            year = int(raw_year)
        elif isinstance(raw_year, str):
            digits = "".join(ch for ch in raw_year if ch.isdigit())
            if len(digits) != 4:
                continue
            year = int(digits)
        else:
            continue
        if isinstance(total, float) and year not in households:
            households[year] = int(round(total * 1000))
    return households


# Pre-1940 decennial census "families" counts (HSUS / decennial census reports).
PRE1940_HOUSEHOLDS: dict[int, int] = {
    1920: 24_351_676,
    1930: 29_904_663,
}


# ---------------------------------------------------------------------------
# Census population
# ---------------------------------------------------------------------------


def load_popclock() -> dict[int, int]:
    text = fetch(SOURCES["census-pop-hist"]["dataFileUrl"], "popclockest.txt").decode("utf-8", "replace")
    pop: dict[int, int] = {}
    for line in text.splitlines():
        m = re.match(r"\s*July 1, (\d{4})\s+([\d,]+)", line)
        if m:
            pop[int(m.group(1))] = int(m.group(2).replace(",", ""))
    return pop


# Decennial census resident-population counts (April 1).
DECENNIAL_POPULATION: dict[int, int] = {
    1920: 106_021_537,
    1930: 123_202_624,
    1940: 132_164_569,
    1950: 151_325_798,
    1960: 179_323_175,
    1970: 203_211_926,
    1980: 226_545_805,
    1990: 248_709_873,
    2000: 281_421_906,
    2010: 308_745_538,
    2020: 331_449_281,
}

# Post-2000 vintage estimates (July 1) for non-census grid years.
MODERN_POPULATION: dict[int, tuple[int, int, bool, str]] = {
    # year -> (value, observation year, interpolated, note)
    2005: (295_516_599, 2005, False, "Census 2000-2010 intercensal national estimate, July 1, 2005."),
    2015: (321_418_820, 2015, False, "Census Vintage 2015 national estimate, July 1, 2015."),
    2025: (
        340_110_988,
        2024,
        True,
        "Latest published Census Vintage 2024 estimate is July 1, 2024; a 2025 figure was not yet available.",
    ),
}


# ---------------------------------------------------------------------------
# Wolff zero/negative wealth (SCF survey years, Table 1 Panel A row 3a)
# ---------------------------------------------------------------------------

WOLFF_SURVEYS: dict[int, float] = {
    1962: 18.2,
    1969: 15.6,
    1983: 15.5,
    1989: 17.9,
    2001: 17.6,
    2007: 18.6,
    2010: 21.8,
    2016: 21.2,
    2019: 19.6,
}


def wolff_cell(year: int) -> dict[str, Any]:
    if year < 1960:
        return na_cell(
            "No reliable primary source for zero/negative net worth before the 1962 SFCC; "
            "metric not reported in estate-tax reconstructions."
        )
    survey_year = min(WOLFF_SURVEYS, key=lambda s: (abs(s - year), s))
    value = WOLFF_SURVEYS[survey_year]
    distance = abs(survey_year - year)
    confidence = "high" if distance <= 1 else "medium" if distance <= 3 else "low"
    return cell(
        value,
        "percent",
        confidence,
        src("wolff", "Table 1 Panel A row 3a"),
        observation_year=survey_year,
        is_interpolated=survey_year != year,
        definition="Percent of all U.S. households with zero or negative net worth",
        methodology_note=(
            f"Survey of Consumer Finances via Wolff; nearest survey year {survey_year}"
            + (f" ({distance} years from grid year)." if distance else ".")
        ),
    )


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------


def population_cell(year: int, popclock: dict[int, int]) -> dict[str, Any]:
    if year in DECENNIAL_POPULATION:
        return cell(
            DECENNIAL_POPULATION[year],
            "count",
            "high",
            src("census-pop-modern", f"{year} decennial census resident population"),
            observation_year=year,
        )
    if year in MODERN_POPULATION:
        value, obs, interp, note = MODERN_POPULATION[year]
        return cell(
            value,
            "count",
            "medium" if interp else "high",
            src("census-pop-modern", f"National population estimate, July 1, {obs}"),
            observation_year=obs,
            is_interpolated=interp,
            methodology_note=note,
        )
    value = popclock[year]
    return cell(
        value,
        "count",
        "high",
        src("census-pop-hist", f"Historical national estimate, July 1, {year}"),
        observation_year=year,
    )


def household_cell(year: int, census_hh: dict[int, int]) -> dict[str, Any]:
    if year in census_hh:
        return cell(
            census_hh[year],
            "count",
            "high",
            src("census-hh", f"Table HH-1 total households, {year}"),
            observation_year=year,
            definition="Total U.S. households (CPS housing-unit definition)",
        )
    if year == 1945:
        # CPS households were not surveyed 1941-1946: interpolate 1940 -> 1947.
        y0, y1 = census_hh[1940], census_hh[1947]
        value = round(y0 + (y1 - y0) * (1945 - 1940) / (1947 - 1940))
        return cell(
            value,
            "count",
            "medium",
            src("census-hh", "Table HH-1, interpolated 1940-1947"),
            observation_year=1945,
            is_interpolated=True,
            definition="Total U.S. households (CPS housing-unit definition)",
            methodology_note="No household survey 1941-1946; linear interpolation between HH-1 values for 1940 and 1947.",
        )
    if year in PRE1940_HOUSEHOLDS:
        return cell(
            PRE1940_HOUSEHOLDS[year],
            "count",
            "medium",
            src("hsus", f"{year} decennial census families count"),
            observation_year=year,
            definition="Census 'families' (private households) count",
            methodology_note="Pre-1940 censuses counted 'families' (private households), a close but not identical concept.",
        )
    # 1925 / 1935: interpolate between adjacent decennial/HH-1 anchors.
    anchors = {**PRE1940_HOUSEHOLDS, 1940: census_hh[1940]}
    prev_year = max(y for y in anchors if y < year)
    next_year = min(y for y in anchors if y > year)
    y0, y1 = anchors[prev_year], anchors[next_year]
    value = round(y0 + (y1 - y0) * (year - prev_year) / (next_year - prev_year))
    return cell(
        value,
        "count",
        "low",
        src("hsus", f"Interpolated {prev_year}-{next_year} decennial counts"),
        observation_year=year,
        is_interpolated=True,
        definition="Census 'families' (private households) count",
        methodology_note=f"Linear interpolation between {prev_year} and {next_year} census counts.",
    )


def build_tier1_row(year: int, shares: dict, levels: dict, popclock: dict[int, int]) -> dict:
    quarter = pick_quarter(shares, year)
    assert quarter, f"No DFA data for {year}"
    share_row = shares[quarter]
    level_row = levels[quarter]
    total_nw = sum(level_row[c]["net_worth"] for c in DFA_CATS)
    total_hh = sum(level_row[c]["household_count"] for c in DFA_CATS)
    dfa_src = src("fed-dfa", f"dfa-networth-*-detail.csv {quarter}")
    share_cells = {}
    for key, cat in zip(SHARE_KEYS, DFA_CATS):
        share_cells[key] = cell(
            round(share_row[cat], 1),
            "percent",
            "high",
            dfa_src,
            observation_year=year,
            definition="Share of aggregate household net worth",
            methodology_note="Household-based DFA; five percentile groups per Fed definitions.",
        )
    return {
        "reportYear": year,
        "tier": 1,
        "totalHouseholdWealth": cell(
            total_nw,
            "usd_millions",
            "high",
            dfa_src,
            observation_year=year,
            definition="Sum of household net worth across DFA percentile groups (nominal USD)",
        ),
        "householdCount": cell(
            total_hh,
            "count",
            "high",
            dfa_src,
            observation_year=year,
            definition="Total households summed across DFA groups",
        ),
        "totalPopulation": population_cell(year, popclock),
        **share_cells,
        "zeroOrNegativeWealth": wolff_cell(year),
    }


def build_tier23_row(
    year: int,
    psz_wealth: dict[int, float],
    psz_shares: dict[int, dict[str, float | None]],
    census_hh: dict[int, int],
    popclock: dict[int, int],
) -> dict:
    tier = 3 if year < 1960 else 2
    year_shares = psz_shares[year]
    share_cells = {}
    for key in SHARE_KEYS:
        value = year_shares[key]
        if value is None:
            share_cells[key] = na_cell(
                "PSZ distributional series reports top-decile detail only before 1962; "
                "no primary bottom-half estimate exists for this year."
            )
            continue
        share_cells[key] = cell(
            value,
            "percent",
            "medium" if year >= 1960 else "low",
            src("psz-dist", "Table TE1, shares of total household wealth"),
            observation_year=year,
            definition="Share of total wealth (equal-split adults, capitalized incomes)",
            methodology_note="PSZ/WID series uses equal-split adults, not identical to post-1989 household DFA.",
        )
    return {
        "reportYear": year,
        "tier": tier,
        "totalHouseholdWealth": cell(
            psz_wealth[year],
            "usd_millions",
            "medium" if year >= 1960 else "low",
            src("psz-agg", "Table TB1 column [1], net household wealth"),
            observation_year=year,
            definition="Total household net worth, nominal USD",
            methodology_note="PSZ national-accounts household series (excludes nonprofits).",
        ),
        "householdCount": household_cell(year, census_hh),
        "totalPopulation": population_cell(year, popclock),
        **share_cells,
        "zeroOrNegativeWealth": wolff_cell(year),
    }


def main() -> None:
    print("Loading Fed DFA...")
    dfa_shares, dfa_levels = load_dfa()
    print("Loading PSZ workbooks...")
    psz_wealth = load_psz_total_wealth()
    psz_shares = load_psz_shares()
    print("Loading Census households and population...")
    census_hh = load_census_households()
    popclock = load_popclock()

    rows = []
    for year in GRID_YEARS:
        if year >= 1990:
            rows.append(build_tier1_row(year, dfa_shares, dfa_levels, popclock))
        else:
            rows.append(build_tier23_row(year, psz_wealth, psz_shares, census_hh, popclock))

    dataset = {
        "version": ACCESS,
        "verifiedAt": ACCESS,
        "methodology": (
            "Every populated cell is reproducible from a cited, downloadable source file. "
            "Tier 1 (1990-2025): Federal Reserve Distributional Financial Accounts for wealth shares, "
            "totals, and household counts; Census for population. "
            "Tier 2 (1960-1985): Piketty-Saez-Zucman appendix tables (equal-split adults) for shares and totals; "
            "Census HH-1 for households. "
            "Tier 3 (1920-1955): PSZ top-decile shares and totals; bottom-half shares and zero-wealth cells are N/A "
            "because no primary source reports them. All dollar amounts are nominal USD."
        ),
        "definitions": {
            "householdWealth": (
                "Household wealth (net worth) is the current market value of all assets owned by households "
                "minus all debts. Assets include financial assets, real estate, business equity, and pension entitlements "
                "(including IRAs and defined-contribution plans), depending on the series. Human capital, future earnings, "
                "and Social Security benefit entitlements are generally excluded. Fed DFA follows Financial Accounts definitions; "
                "historical PSZ series capitalize taxable capital incomes and use equal-split adults."
            ),
            "household": (
                "A household is a person or group of people who occupy a housing unit. Census and Survey of Consumer Finances "
                "use this housing-unit definition. Federal Reserve DFA distributes aggregate household wealth across percentile groups "
                "of the household wealth distribution. Historical PSZ/WID series often use equal-split adults (wealth split equally "
                "between spouses) or tax units, which can differ slightly from household counts."
            ),
        },
        "sources": [SOURCES[k] for k in sorted(SOURCES)],
        "rows": rows,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(dataset, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
